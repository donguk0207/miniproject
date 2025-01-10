from flask import Flask, render_template, request, redirect, url_for, session
from openpyxl import load_workbook
from db_connector import fetch_data_from_mysql
from map_generator import create_map
import secrets
import apt
import graph_generation
from graph_data import graph_data
import threading

app = Flask(__name__)
app.secret_key = secrets.token_hex(16)

# 로그인 페이지
@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        try:
            wb = load_workbook('login.xlsx')
            sheet = wb.active
            users = [row for row in sheet.iter_rows(min_row=2, values_only=True)]

            for user in users:
                if user[0] == username and user[1] == password:
                    session['username'] = username
                    return redirect(url_for('dashboard'))

            return 'Invalid login credentials'
        except Exception as e:
            return f"Error: {e}"

    return render_template('index.html')

@app.route('/dashboard', methods=['GET'])
def dashboard():
    if 'username' not in session:
        return redirect(url_for('login'))

    return render_template('dashboard.html', username=session['username'])

# 지도 페이지
@app.route('/map', methods=['GET'])
def map():
    if 'username' not in session:
        return redirect(url_for('login'))

    # MySQL에서 데이터 가져오기
    data = fetch_data_from_mysql()
    create_map(data)

    return render_template('map.html')

@app.route('/graph', methods=['GET'])
def graph():
    # 그래프 이미지 경로 얻기
    graph_image_path = graph_data()

    # 그래프 이미지를 렌더링하는 새로운 페이지로 이동
    return render_template('graph_page.html', graph_image_path=graph_image_path)

# 정보 페이지
@app.route('/information', methods=['GET'])
def information():
    if 'username' not in session:
        return redirect(url_for('login'))

    graph_generation.generate_graph()  # 그래프 생성
    graph_generation.generate_seocho_graph()
    return render_template('information.html')  #

# Tkinter 프로그램을 실행하는 라우터
@app.route('/run_tkinter', methods=['GET'])
def run_tkinter():
    if 'username' not in session:
        return redirect(url_for('login'))

    def run_tkinter_thread():
        scraper = apt.NaverRealEstateScraper()
        scraper.run_tkinter_program()  # Tkinter 프로그램 실행

    threading.Thread(target=run_tkinter_thread, daemon=True).start()

    return redirect(url_for('dashboard'))

if __name__ == "__main__":
    app.run(debug=True)